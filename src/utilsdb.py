
import pandas as pd
import logging
from typing import Dict, Any, Optional


# En Docker, la ruta debe ser absoluta
#db_path = '/app/database/mecanicos.db'

import os


def initialize_database_mysql():
    """
    Initializes the MySQL database with the provided schema.
    """
    try:
        # Conectar al servidor MySQL y a la base de datos 'mecanicos'
        connection = mysql.connector.connect(
            host='db',  # Cambia según tu configuración
            port=3306,  # Puerto de conexión
            user='mecanicos_user',  # Usuario de MySQL definido en Docker
            password='mecanicos_password',  # Contraseña de MySQL definida en Docker
            database='mecanicos'  # Conectar directamente a la base de datos 'mecanicos'
        )

        if connection.is_connected():
            cursor = connection.cursor()

            # Script SQL para crear las tablas
            sql_script = """
                -- Table: Productos
                CREATE TABLE IF NOT EXISTS Productos (
                    ID INT PRIMARY KEY NOT NULL AUTO_INCREMENT,
                    Nombre VARCHAR(255) NOT NULL,
                    Categoria VARCHAR(255) NOT NULL
                );

                -- Set the starting value of AUTO_INCREMENT to 1 every time the table is recreated or modified
                ALTER TABLE Productos AUTO_INCREMENT = 1;

                -- Table: Orden_Entrada
                CREATE TABLE IF NOT EXISTS Orden_Entrada (
                    ID_Entrada INT PRIMARY KEY AUTO_INCREMENT,
                    Fecha_Entrada DATE NOT NULL,
                    Status ENUM('liberada', 'proceso', 'inactiva') NOT NULL,
                    Fecha_Salida DATE,
                    ID_Camion VARCHAR(255),
                    Motivo_Entrada TEXT,
                    Motivo_Salida TEXT,
                    Tipo ENUM('CONSUMIBLE', 'PREVENTIVO', 'CORRECTIVO'),
                    Kilometraje INT,
                    Lugar ENUM('PUEBLA', 'VILLA HERMOSA', 'GUADALAJARA'),
                    hora_registro TIME,
                    hora_salida TIME
                );

                -- Table: Camion
                CREATE TABLE IF NOT EXISTS Camion (
                    VIN VARCHAR(255) PRIMARY KEY,
                    NumeroUnidad INT NOT NULL,
                    Kilometraje INT NOT NULL,
                    Marca VARCHAR(255) NOT NULL,
                    Modelo VARCHAR(255) NOT NULL
                );

                -- Table: Productos_Servicio
                CREATE TABLE IF NOT EXISTS Productos_Servicio (
                    ID_Orden INT NOT NULL,
                    ID_Producto INT NOT NULL,
                    Cantidad INT NOT NULL,
                    PRIMARY KEY (ID_Orden, ID_Producto),
                    FOREIGN KEY (ID_Orden) REFERENCES Orden_Entrada(ID_Entrada),
                    FOREIGN KEY (ID_Producto) REFERENCES Productos(ID)
                );
            """

            # Ejecutar el script SQL
            for result in cursor.execute(sql_script, multi=True):
                pass  # Consume el resultado para mantener la conexión sincronizada

            # Confirmar los cambios
            connection.commit()

            print("Database 'mecanicos' initialized successfully!")

    except Error as e:
        print(f"An error occurred while initializing the database: {e}")
        raise e

    except Exception as ex:
        print(f"An unexpected error occurred: {ex}")
        raise ex

    finally:
        # Cerrar la conexión
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()


#Camion
def create_camion(connection, vin, numero_unidad, kilometraje, marca, modelo):
    try:
        cursor = connection.cursor()
        query = "INSERT INTO Camion (VIN, NumeroUnidad, Kilometraje, Marca, Modelo) VALUES (%s, %s, %s, %s, %s)"
        cursor.execute(query, (vin, numero_unidad, kilometraje, marca, modelo))
        connection.commit()
        print("Camion created successfully!")
    except Error as e:
        print(f"An error occurred while creating Camion: {e}")
        raise e
    finally:
        if connection.is_connected():
            cursor.close()

def read_camion(connection, vin):
    try:
        cursor = connection.cursor()
        query = "SELECT * FROM Camion WHERE VIN = %s"
        cursor.execute(query, (vin,))
        return cursor.fetchone()
    except Error as e:
        print(f"An error occurred while reading Camion: {e}")
        raise e
    finally:
        if connection.is_connected():
            cursor.close()


import mysql.connector
from mysql.connector import Error

# Update Camion
def update_camion(connection, vin, numero_unidad=None, kilometraje=None, marca=None, modelo=None):
    try:
        cursor = connection.cursor()
        fields = []
        values = []

        if numero_unidad is not None:
            fields.append("NumeroUnidad = %s")
            values.append(numero_unidad)
        if kilometraje is not None:
            fields.append("Kilometraje = %s")
            values.append(kilometraje)
        if marca is not None:
            fields.append("Marca = %s")
            values.append(marca)
        if modelo is not None:
            fields.append("Modelo = %s")
            values.append(modelo)

        if not fields:
            return {"error": "No fields to update."}

        values.append(vin)
        query = f"UPDATE Camion SET {', '.join(fields)} WHERE VIN = %s"
        cursor.execute(query, values)
        connection.commit()

        return {"message": "Camion updated successfully."}
    except Error as e:
        print(f"An error occurred while updating Camion: {e}")
        raise e
    finally:
        if connection.is_connected():
            cursor.close()

# Delete Camion
def delete_camion(connection, vin):
    try:
        cursor = connection.cursor()
        query = "DELETE FROM Camion WHERE VIN = %s"
        cursor.execute(query, (vin,))
        connection.commit()
        print("Camion deleted successfully!")
    except Error as e:
        print(f"An error occurred while deleting Camion: {e}")
        raise e
    finally:
        if connection.is_connected():
            cursor.close()

# Create Productos_Servicio
def create_productos_servicio(connection, id_orden, id_producto, cantidad):
    try:
        cursor = connection.cursor()
        # Ensure the product exists
        cursor.execute("SELECT ID FROM Productos WHERE ID = %s", (id_producto,))
        if not cursor.fetchone():
            return {"error": f"El producto con ID {id_producto} no existe."}

        # Check if the product is already associated with the order
        cursor.execute("SELECT Cantidad FROM Productos_Servicio WHERE ID_Orden = %s AND ID_Producto = %s", (id_orden, id_producto))
        existing_record = cursor.fetchone()
        
        if existing_record:
            # Update the quantity if the product is already associated with the order
            nueva_cantidad = existing_record[0] + cantidad
            update_query = "UPDATE Productos_Servicio SET Cantidad = %s WHERE ID_Orden = %s AND ID_Producto = %s"
            cursor.execute(update_query, (nueva_cantidad, id_orden, id_producto))
            connection.commit()
            return {"message": f"Cantidad del producto con ID {id_producto} actualizada en la orden ID {id_orden} (Nueva cantidad: {nueva_cantidad})."}
        else:
            # Insert into Productos_Servicio if not already associated
            insert_query = "INSERT INTO Productos_Servicio (ID_Orden, ID_Producto, Cantidad) VALUES (%s, %s, %s)"
            cursor.execute(insert_query, (id_orden, id_producto, cantidad))
            connection.commit()
            return {"message": f"El producto con ID {id_producto} se ha asociado con la orden ID {id_orden} (Cantidad: {cantidad})."}
    except Error as e:
        return {"error": f"Ocurrió un error: {e}"}
    finally:
        if connection.is_connected():
            cursor.close()

# Read Productos_Servicio
def read_productos_servicio(connection, id_orden):
    try:
        cursor = connection.cursor()
        query = """
        SELECT ps.ID_Orden, ps.ID_Producto, p.Nombre, p.Categoria, ps.Cantidad
        FROM Productos_Servicio ps
        JOIN Productos p ON ps.ID_Producto = p.ID
        WHERE ps.ID_Orden = %s
        """
        cursor.execute(query, (id_orden,))
        results = cursor.fetchall()
        return results
    except Error as e:
        print(f"An error occurred while reading Productos_Servicio: {e}")
        raise e
    finally:
        if connection.is_connected():
            cursor.close()

# Delete Productos_Servicio
def delete_productos_servicio(connection, id_orden, id_producto):
    try:
        cursor = connection.cursor()
        query = "DELETE FROM Productos_Servicio WHERE ID_Orden = %s AND ID_Producto = %s"
        cursor.execute(query, (id_orden, id_producto))
        connection.commit()
        return {"message": "Product removed from service successfully."}
    except Error as e:
        print(f"An error occurred while deleting Productos_Servicio: {e}")
        raise e
    finally:
        if connection.is_connected():
            cursor.close()

# Update Productos_Servicio
def update_productos_servicio(connection, id_orden, id_producto, new_id_orden=None, new_id_producto=None):
    try:
        cursor = connection.cursor()
        fields = []
        values = []

        if new_id_orden is not None:
            fields.append("ID_Orden = %s")
            values.append(new_id_orden)
        if new_id_producto is not None:
            fields.append("ID_Producto = %s")
            values.append(new_id_producto)

        if not fields:
            return {"error": "No fields to update."}

        values.append(id_orden)
        values.append(id_producto)
        query = f"UPDATE Productos_Servicio SET {', '.join(fields)} WHERE ID_Orden = %s AND ID_Producto = %s"

        cursor.execute(query, values)
        connection.commit()
        return {"message": "Product-service association updated successfully."}
    except Error as e:
        return {"error": f"An error occurred: {e}"}
    finally:
        if connection.is_connected():
            cursor.close()

    
#Orden_Entrada
import mysql.connector
from mysql.connector import Error, IntegrityError
from typing import Dict, Any, Optional
import logging

# Create Orden_Entrada
def create_orden_entrada(
    connection,
    fecha_entrada: str,
    status: str,
    id_camion: str,
    motivo_entrada: str,
    tipo: str,
    hora_registro: str,
    lugar: str,
    fecha_salida: Optional[str] = None,
    kilometraje: Optional[int] = None,
    motivo_salida: Optional[str] = None
) -> Dict[str, Any]:
    """
    Creates a new 'Orden_Entrada' record in the database.

    Args:
        connection: Database connection.
        fecha_entrada (str): Entry date in 'YYYY-MM-DD' format.
        status (str): Initial status of the order. Must be one of: 'liberada', 'proceso', 'inactiva'.
        id_camion (str): Truck VIN.
        motivo_entrada (str): Reason for entry.
        tipo (str): Type of maintenance ('CONSUMIBLE', 'PREVENTIVO', 'CORRECTIVO').
        hora_registro (str): Entry time in 'HH:MM' format.
        lugar (str): Location of the maintenance. ('PUEBLA', 'VILLA HERMOSA', 'GUADALAJARA').
        fecha_salida (str, optional): Exit date in 'YYYY-MM-DD' format.
        kilometraje (int, optional): Exit mileage.
        motivo_salida (str, optional): Reason for exit.

    Returns:
        dict: Success message or error message.
    """
    # Validate the status before attempting insertion
    VALID_STATUS = {'liberada', 'proceso', 'inactiva'}
    status = status.lower()  # Convert to lowercase for consistency

    if status not in VALID_STATUS:
        return {
            "error": f"Invalid status. Must be one of: {', '.join(VALID_STATUS)}"
        }

    # Validate the type of maintenance
    VALID_TIPOS = {'CONSUMIBLE', 'PREVENTIVO', 'CORRECTIVO'}
    tipo = tipo.upper()  # Convert to uppercase for consistency

    if tipo not in VALID_TIPOS:
        return {
            "error": f"Invalid type. Must be one of: {', '.join(VALID_TIPOS)}"
        }

    # Validate the place
    VALID_LUGARES = {'PUEBLA', 'VILLA HERMOSA', 'GUADALAJARA'}
    lugar = lugar.upper()  # Convert to uppercase for consistency

    if lugar not in VALID_LUGARES:
        return {
            "error": f"Invalid place. Must be one of: {', '.join(VALID_LUGARES)}"
        }

    cursor = connection.cursor()
    try:
        query = """
        INSERT INTO Orden_Entrada (
            Fecha_Entrada,
            Status,
            Fecha_Salida,
            ID_Camion,
            Motivo_Entrada,
            Motivo_Salida,
            Tipo,
            Kilometraje,
            Lugar,
            Hora_Registro
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        values = (
            fecha_entrada,
            status,
            fecha_salida,
            id_camion,
            motivo_entrada,
            motivo_salida,
            tipo,
            kilometraje,
            lugar,
            hora_registro
        )

        logging.debug(f"Executing query with values: {values}")
        cursor.execute(query, values)
        connection.commit()

        return {
            "message": "The entry order has been successfully created.",
            "status": status,
            "tipo": tipo
        }

    except IntegrityError as e:
        logging.error(f"Integrity error creating Orden_Entrada: {e}")
        connection.rollback()
        return {
            "error": f"An integrity error occurred while creating the order: {str(e)}",
            "details": {
                "status_provided": status,
                "tipo_provided": tipo,
                "lugar_provided": lugar
            }
        }
    except Error as e:
        logging.error(f"Error creating Orden_Entrada: {e}")
        connection.rollback()
        return {
            "error": f"An error occurred while creating the order: {str(e)}",
            "details": {
                "status_provided": status,
                "tipo_provided": tipo,
                "lugar_provided": lugar
            }
        }
    finally:
        cursor.close()

# Salida Orden_Entrada
def salida_orden_entrada(conn, orden_id: int, fecha_salida: str, status: str, motivo_salida: str, kilometraje: int, hora_salida: str):
    """
    Actualiza 'fecha_salida', 'status' y 'motivo_salida' de una orden existente en Orden_Entrada.
    
    Args:
        conn: Conexión a la base de datos.
        orden_id (int): ID de la orden a actualizar.
        fecha_salida (str): Fecha de salida en formato 'YYYY-MM-DD'.
        status (str): Nuevo estado de la orden.
        motivo_salida (str): Motivo de la salida.
        kilometraje (int): Kilometraje al salir.
        hora_salida (str): Hora de salida en formato 'HH:MM'.
    
    Returns:
        dict: Mensaje de éxito o error.
    """
    cursor = conn.cursor()
    try:
        # Verificar si la orden existe
        cursor.execute("SELECT * FROM Orden_Entrada WHERE ID_Entrada = %s", (orden_id,))
        orden = cursor.fetchone()
        if not orden:
            return {"error": f"No se encontró una orden con ID {orden_id}."}

        # Actualizar la orden con los nuevos datos
        query = """
        UPDATE Orden_Entrada
        SET Fecha_Salida = %s,
            Status = %s,
            Motivo_Salida = %s,
            Kilometraje = %s,
            Hora_Salida = %s
        WHERE ID_Entrada = %s
        """
        cursor.execute(query, (
            fecha_salida,   # Fecha_Salida
            status,         # Status
            motivo_salida,  # Motivo_Salida
            kilometraje,    # Kilometraje
            hora_salida,    # Hora_Salida
            orden_id        # WHERE ID_Entrada
        ))
        conn.commit()
        return {"message": f"La orden con ID {orden_id} ha sido actualizada con éxito."}
    except Exception as e:
        conn.rollback()  # Revertir cambios en caso de error
        return {"error": f"Ocurrió un error al actualizar la orden: {e}"}
    finally:
        cursor.close()


# Read Orden_Entrada
def read_orden_entrada(connection, orden_id):
    cursor = connection.cursor()
    try:
        query = "SELECT * FROM Orden_Entrada WHERE ID_Entrada = %s"
        logging.debug(f"Executing query: {query} with orden_id={orden_id}")
        cursor.execute(query, (orden_id,))
        result = cursor.fetchone()
        return result
    except Error as e:
        logging.error(f"Error reading Orden_Entrada: {e}")
        raise e
    finally:
        cursor.close()

# Update Orden_Entrada
def update_orden_entrada(connection, orden_id, fecha_entrada=None, status=None, fecha_salida=None, id_camion=None, motivo_entrada=None, motivo_salida=None, tipo=None, kilometraje=None):
    cursor = connection.cursor()
    try:
        # Update only the fields that are provided
        fields = []
        values = []

        if fecha_entrada is not None:
            fields.append("Fecha_Entrada = %s")
            values.append(fecha_entrada)
        if status is not None:
            fields.append("Status = %s")
            values.append(status)
        if fecha_salida is not None:
            fields.append("Fecha_Salida = %s")
            values.append(fecha_salida)
        if id_camion is not None:
            fields.append("ID_Camion = %s")
            values.append(id_camion)
        if motivo_entrada is not None:
            fields.append("Motivo_Entrada = %s")
            values.append(motivo_entrada)
        if motivo_salida is not None:
            fields.append("Motivo_Salida = %s")
            values.append(motivo_salida)
        if tipo is not None:
            fields.append("Tipo = %s")
            values.append(tipo)
        if kilometraje is not None:
            fields.append("Kilometraje = %s")
            values.append(kilometraje)

        if not fields:
            return {"error": "No fields to update."}

        values.append(orden_id)
        query = f"UPDATE Orden_Entrada SET {', '.join(fields)} WHERE ID_Entrada = %s"
        cursor.execute(query, values)
        connection.commit()
        return {"message": "Orden_Entrada updated successfully."}
    except Error as e:
        return {"error": f"An error occurred while updating the order: {e}"}
    finally:
        cursor.close()

# Delete Orden_Entrada
def delete_orden_entrada(connection, orden_id):
    cursor = connection.cursor()
    try:
        # Check if the record exists
        query_check = "SELECT * FROM Orden_Entrada WHERE ID_Entrada = %s"
        cursor.execute(query_check, (orden_id,))
        record = cursor.fetchone()

        if not record:
            return {"message": f"No record found with ID: {orden_id}"}

        # Delete the record
        query_delete = "DELETE FROM Orden_Entrada WHERE ID_Entrada = %s"
        cursor.execute(query_delete, (orden_id,))
        connection.commit()

        return {"message": f"Orden_Entrada with ID: {orden_id} deleted successfully."}
    except Error as e:
        return {"error": f"An error occurred while deleting the order: {e}"}
    finally:
        cursor.close()

#Producto
import mysql.connector
from mysql.connector import Error, IntegrityError
import logging
from typing import Dict, Any, Optional

# Create Producto
def create_producto(connection, nombre: str, cantidad: int, categoria: str):
    cursor = connection.cursor()
    try:
        query = "INSERT INTO Productos (Nombre, Categoria) VALUES (%s, %s)"
        cursor.execute(query, (nombre, categoria))
        connection.commit()
        producto_id = cursor.lastrowid
        return {"message": "Producto created successfully.", "producto_id": producto_id}
    except Error as e:
        logging.error(f"An error occurred while creating Producto: {e}")
        return {"error": f"An error occurred while creating Producto: {e}"}
    finally:
        cursor.close()

# Read Producto
def read_producto(connection, producto_id: int):
    cursor = connection.cursor()
    try:
        query = "SELECT * FROM Productos WHERE ID = %s"
        cursor.execute(query, (producto_id,))
        return cursor.fetchone()
    except Error as e:
        logging.error(f"An error occurred while reading Producto: {e}")
        raise e
    finally:
        cursor.close()

# Update Producto
def update_producto(connection, producto_id: int, nombre: Optional[str]=None, cantidad: Optional[int]=None, categoria: Optional[str]=None):
    cursor = connection.cursor()
    try:
        fields = []
        values = []

        if nombre is not None:
            fields.append("Nombre = %s")
            values.append(nombre)
        if cantidad is not None:
            fields.append("Cantidad = %s")
            values.append(cantidad)
        if categoria is not None:
            fields.append("Categoria = %s")
            values.append(categoria)

        if not fields:
            return {"error": "No fields to update."}

        values.append(producto_id)
        query = f"UPDATE Productos SET {', '.join(fields)} WHERE ID = %s"
        cursor.execute(query, values)
        connection.commit()

        return {"message": "Producto updated successfully."}
    except Error as e:
        logging.error(f"An error occurred while updating Producto: {e}")
        return {"error": f"An error occurred while updating Producto: {e}"}
    finally:
        cursor.close()

# Delete Producto
def delete_producto(connection, producto_id: int):
    cursor = connection.cursor()
    try:
        query = "DELETE FROM Productos WHERE ID = %s"
        cursor.execute(query, (producto_id,))
        connection.commit()
        return {"message": "Producto deleted successfully."}
    except Error as e:
        logging.error(f"An error occurred while deleting Producto: {e}")
        return {"error": f"An error occurred while deleting Producto: {e}"}
    finally:
        cursor.close()

# Add Products for Order
def add_products_for_order(conn, orden_id: int, product_details):
    """
    Agrega productos a una orden en la base de datos, validando cada producto y cantidad.
    
    Args:
        conn: Conexión a la base de datos.
        orden_id (int): ID de la orden a la que se agregan los productos.
        product_details (list): Lista de diccionarios con detalles de productos (id_producto y cantidad).
    
    Returns:
        list: Mensajes de éxito o error para cada producto.
    """
    cursor = conn.cursor()
    messages = []
    try:
        for product_detail in product_details:
            id_producto = product_detail.get('id_producto')
            cantidad = product_detail.get('cantidad')

            if id_producto is None or cantidad is None:
                messages.append({"error": "Faltan detalles del producto (ID o cantidad)."})
                continue

            try:
                id_producto = int(id_producto)
                cantidad = int(cantidad)

                # Verificar que el producto existe
                cursor.execute("SELECT ID, Nombre, Categoria FROM Productos WHERE ID = %s", (id_producto,))
                product = cursor.fetchone()
                if not product:
                    messages.append({"error": f"El producto con ID {id_producto} no existe."})
                    continue

                if cantidad <= 0:
                    messages.append({"error": "La cantidad debe ser un entero positivo."})
                    continue

                # Asociar el producto con la orden mediante su ID
                result = create_productos_servicio(conn, orden_id, id_producto, cantidad)
                if 'error' in result:
                    messages.append(result)
                else:
                    messages.append({"success": f"Producto {product[1]} (ID: {id_producto}) agregado correctamente a la orden {orden_id}."})
            except ValueError:
                messages.append({"error": "ID de producto o cantidad inválidos."})
    except Exception as e:
        conn.rollback()  # Revertir cambios en caso de error global
        messages.append({"error": f"Ocurrió un error al agregar productos a la orden: {e}"})
    finally:
        cursor.close()
    return messages


#Create Products

import mysql.connector
from mysql.connector import Error
import logging
from typing import Dict, Any, Optional

# Clear Productos table
def clear_productos_table(connection):
    """Deletes all records from the Productos table."""
    cursor = connection.cursor()
    try:
        cursor.execute("DELETE FROM Productos")
        connection.commit()
        print("All records deleted from Productos table.")
    except Error as e:
        print(f"An error occurred while clearing Productos table: {e}")
    finally:
        cursor.close()

# Add categories and products
def add_categories_and_products(connection, categories_and_products):
    """
    Adds categories and their associated products to the Productos table.
    Args:
        connection: Database connection.
        categories_and_products (list): List of dictionaries with 'category' and 'products' keys.
    """
    cursor = connection.cursor()
    try:
        # Eliminar todos los registros de Productos y reiniciar el AUTO_INCREMENT
        cursor.execute("DELETE FROM Productos;")
        cursor.execute("ALTER TABLE Productos AUTO_INCREMENT = 1;")

        for category_data in categories_and_products:
            category = category_data['category']
            products = category_data['products']
            for product_name in products:
                # Clean up product name (strip whitespace)
                product_name = product_name.strip()
                # Insert product into Productos table without predefined ID
                query = "INSERT INTO Productos (Nombre, Categoria) VALUES (%s, %s)"
                cursor.execute(query, (product_name, category))
        
        connection.commit()
        print("Categories and products added successfully.")
    except Error as e:
        print(f"An error occurred while adding categories and products: {e}")
    finally:
        cursor.close()

# Information queries
def get_products_used_in_order(connection, id_orden):
    """
    Recupera y devuelve una lista de productos utilizados en una orden específica en formato de texto legible.
    Args:
        connection: Conexión a la base de datos.
        id_orden (int): ID de la orden.
    Returns:
        str: Una cadena de texto con los productos utilizados o un mensaje indicando que no se encontraron productos.
    """
    cursor = connection.cursor()
    try:
        query = """
        SELECT p.Nombre, p.Categoria, ps.Cantidad
        FROM Productos_Servicio ps
        JOIN Productos p ON ps.ID_Producto = p.ID
        WHERE ps.ID_Orden = %s
        """
        cursor.execute(query, (id_orden,))
        results = cursor.fetchall()

        if not results:
            return f"No se encontraron productos para la orden con ID {id_orden}."

        response = f"Productos utilizados en la orden {id_orden}:\n"
        for nombre, categoria, cantidad in results:
            response += f"- {nombre} (Categoría: {categoria}), Cantidad: {cantidad}\n"
        return response
    except Error as e:
        print(f"An error occurred while retrieving products used in order: {e}")
        return f"Ocurrió un error: {e}"
    finally:
        cursor.close()

def get_products_used_in_month(connection, year: int, month: int):
    """
    Recupera todos los productos utilizados y sus cantidades en todas las órdenes durante un mes específico, y los devuelve en un formato de texto legible.
    Args:
        connection: Conexión a la base de datos.
        year (int): Año de interés.
        month (int): Mes de interés (1-12).
    Returns:
        str: Una cadena de texto con los productos utilizados o un mensaje indicando que no se encontraron productos.
    """
    cursor = connection.cursor()
    try:
        query = """
        SELECT p.Nombre, p.Categoria, SUM(ps.Cantidad) as TotalCantidad
        FROM Productos_Servicio ps
        JOIN Productos p ON ps.ID_Producto = p.ID
        JOIN Orden_Entrada oe ON ps.ID_Orden = oe.ID_Entrada
        WHERE YEAR(oe.Fecha_Entrada) = %s AND MONTH(oe.Fecha_Entrada) = %s
        GROUP BY p.ID
        ORDER BY TotalCantidad DESC
        """
        cursor.execute(query, (year, month))
        results = cursor.fetchall()

        if not results:
            return f"No se encontraron productos utilizados en {month}/{year}."

        response = f"Productos utilizados en {month}/{year}:\n"
        for nombre, categoria, total_cantidad in results:
            response += f"- {nombre} (Categoría: {categoria}), Cantidad total: {total_cantidad}\n"
        return response
    except Error as e:
        print(f"An error occurred while retrieving products used in month: {e}")
        return f"Ocurrió un error: {e}"
    finally:
        cursor.close()

def get_product_usage_for_truck(connection, vin: str, product_id: int, start_date: str, end_date: str):
    """
    Recupera la cantidad total de un producto específico utilizado en un camión determinado dentro de un rango de fechas, y la devuelve en formato de texto legible.
    Args:
        connection: Conexión a la base de datos.
        vin (str): VIN del camión.
        product_id (int): ID del producto.
        start_date (str): Fecha de inicio en formato 'YYYY-MM-DD'.
        end_date (str): Fecha de fin en formato 'YYYY-MM-DD'.
    Returns:
        str: Una cadena de texto con la cantidad total utilizada o un mensaje indicando que no se encontró uso del producto.
    """
    cursor = connection.cursor()
    try:
        query = """
        SELECT p.Nombre, p.Categoria, SUM(ps.Cantidad) as TotalCantidad
        FROM Productos_Servicio ps
        JOIN Orden_Entrada oe ON ps.ID_Orden = oe.ID_Entrada
        JOIN Productos p ON ps.ID_Producto = p.ID
        WHERE oe.ID_Camion = %s AND ps.ID_Producto = %s
          AND oe.Fecha_Entrada BETWEEN %s AND %s
        GROUP BY p.ID
        """
        cursor.execute(query, (vin, product_id, start_date, end_date))
        result = cursor.fetchone()

        if result and result[2]:
            nombre = result[0]
            categoria = result[1]
            total_cantidad = result[2]
            return (f"El producto '{nombre}' (Categoría: {categoria}) se utilizó {total_cantidad} "
                    f"veces para el camión con VIN {vin} entre {start_date} y {end_date}.")
        else:
            return (f"No se encontró uso del producto con ID {product_id} para el camión con VIN {vin} "
                    f"entre {start_date} y {end_date}.")
    except Error as e:
        print(f"An error occurred while retrieving product usage for truck: {e}")
        return f"Ocurrió un error: {e}"
    finally:
        cursor.close()

def get_products_used_for_truck(connection, vin: str):
    """
    Recupera todos los productos utilizados y sus cantidades para un camión específico, y los devuelve en un formato de texto legible.
    Args:
        connection: Conexión a la base de datos.
        vin (str): VIN del camión.
    Returns:
        str: Una cadena de texto con los productos utilizados o un mensaje indicando que no se encontraron productos.
    """
    cursor = connection.cursor()
    try:
        query = """
        SELECT p.Nombre, p.Categoria, SUM(ps.Cantidad) as TotalCantidad
        FROM Productos_Servicio ps
        JOIN Productos p ON ps.ID_Producto = p.ID
        JOIN Orden_Entrada oe ON ps.ID_Orden = oe.ID_Entrada
        WHERE oe.ID_Camion = %s
        GROUP BY p.ID
        ORDER BY TotalCantidad DESC
        """
        cursor.execute(query, (vin,))
        results = cursor.fetchall()

        if not results:
            return f"No se encontraron productos utilizados para el camión con VIN {vin}."

        response = f"Productos utilizados para el camión con VIN {vin}:\n"
        for nombre, categoria, total_cantidad in results:
            response += f"- {nombre} (Categoría: {categoria}), Cantidad total: {total_cantidad}\n"
        return response
    except Error as e:
        print(f"An error occurred while retrieving products used for truck: {e}")
        return f"Ocurrió un error: {e}"
    finally:
        cursor.close()

def get_order_count_for_branch(connection, sucursal: str):
    """
    Recupera el número de órdenes para una sucursal específica y lo devuelve en formato de texto legible.
    Args:
        connection: Conexión a la base de datos.
        sucursal (str): Nombre o identificador de la sucursal.
    Returns:
        str: Una cadena de texto con el número de órdenes o un mensaje indicando que no se encontraron órdenes.
    """
    cursor = connection.cursor()
    try:
        query = """
        SELECT COUNT(*) as OrderCount
        FROM Orden_Entrada
        WHERE Lugar = %s
        """
        cursor.execute(query, (sucursal,))
        result = cursor.fetchone()
        count = result[0] if result else 0

        if count > 0:
            return f"La sucursal '{sucursal}' ha procesado {count} órdenes."
        else:
            return f"No se encontraron órdenes para la sucursal '{sucursal}'."
    except Error as e:
        print(f"An error occurred while retrieving order count for branch: {e}")
        return f"Ocurrió un error: {e}"
    finally:
        cursor.close()

def get_order_details_for_truck(connection, vin: str):
    """
    Recupera las órdenes, motivos y fechas para un camión específico y las devuelve en un formato de texto legible.
    Args:
        connection: Conexión a la base de datos.
        vin (str): VIN del camión.
    Returns:
        str: Una cadena de texto con los detalles de las órdenes o un mensaje indicando que no se encontraron órdenes.
    """
    cursor = connection.cursor()
    try:
        query = """
        SELECT ID_Entrada, Motivo_Entrada, Fecha_Entrada
        FROM Orden_Entrada
        WHERE ID_Camion = %s
        ORDER BY Fecha_Entrada DESC
        """
        cursor.execute(query, (vin,))
        orders = cursor.fetchall()

        if not orders:
            return f"No se encontraron órdenes para el camión con VIN {vin}."

        response = f"Órdenes para el camión con VIN {vin}:\n"
        for orden_id, motivo, fecha_entrada in orders:
            response += f"- Orden ID: {orden_id}, Motivo: {motivo}, Fecha de Entrada: {fecha_entrada}\n"
        return response
    except Error as e:
        print(f"An error occurred while retrieving order details for truck: {e}")
        return f"Ocurrió un error: {e}"
    finally:
        cursor.close()


#Excel Table

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension

def export_orders_to_excel(connection, excel_path: str):
    """
    Exports specific fields from the database to an Excel file with separate sheets for each 'Lugar'.
    Creates the file and its directory if they do not exist.

    Args:
        connection: MySQL database connection.
        excel_path (str): Path where the Excel file will be saved.

    Returns:
        None
    """
    try:
        # Ensure that the directory for the Excel file exists
        directory = os.path.dirname(excel_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
            print(f"Directory '{directory}' created to store the Excel file.")

        # Fetch distinct values of Lugar
        cursor = connection.cursor()
        cursor.execute("SELECT DISTINCT Lugar FROM Orden_Entrada")
        distinct_places = [row[0] for row in cursor.fetchall()]
        cursor.close()

        # Create an Excel workbook
        wb = Workbook()
        # Remove the default sheet created automatically
        default_sheet = wb.active
        wb.remove(default_sheet)

        for place in distinct_places:
            # Define the SQL query to filter by Lugar
            query = """
            SELECT 
                oe.ID_Entrada AS `ID Orden`,
                DATE_FORMAT(oe.Fecha_Entrada, '%%d/%%m/%%Y') AS `Fecha de Entrada`,
                c.VIN AS `VIN`,
                c.Modelo AS `Modelo`,
                c.Marca AS `Marca`,
                oe.Motivo_Entrada AS `Motivo de Entrada`,
                oe.Tipo AS `Tipo de Mantenimiento`,
                oe.Status AS `Estado`,
                GROUP_CONCAT(DISTINCT p.Categoria SEPARATOR ', ') AS `Categorías`,
                GROUP_CONCAT(CONCAT(IFNULL(p.Nombre, 'Sin productos'), ' x', IFNULL(ps.Cantidad, ''))) AS `Productos y Cantidades`,
                IFNULL(c.Kilometraje, 0) AS `Kilometraje`,
                oe.Motivo_Salida AS `Motivo de Salida`,
                IFNULL(DATE_FORMAT(oe.Fecha_Salida, '%%d/%%m/%%Y'), 'Pendiente') AS `Fecha de Salida`
            FROM Orden_Entrada oe
            LEFT JOIN Camion c ON oe.ID_Camion = c.VIN
            LEFT JOIN Productos_Servicio ps ON oe.ID_Entrada = ps.ID_Orden
            LEFT JOIN Productos p ON ps.ID_Producto = p.ID
            WHERE oe.Lugar = %s
            GROUP BY oe.ID_Entrada
            ORDER BY oe.Fecha_Entrada DESC;
            """

            # Execute the query and load the data into a pandas DataFrame
            df = pd.read_sql(query, connection, params=(place,))

            # Verify if the DataFrame is empty
            if df.empty:
                print(f"No data found for Lugar '{place}' to export.")
                continue

            # Create a new sheet for each 'Lugar'
            ws = wb.create_sheet(title=place[:31])  # Excel sheet names must be 31 characters or fewer

            # Add a title to the worksheet
            ws.merge_cells('A1:M1')
            title_cell = ws['A1']
            title_cell.value = f"Ordenes de Entrada - {place}"
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Append the DataFrame to the worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                ws.append(row)

            # Set styles for the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border_style = Border(left=Side(style='thin', color='000000'),
                                  right=Side(style='thin', color='000000'),
                                  top=Side(style='thin', color='000000'),
                                  bottom=Side(style='thin', color='000000'))

            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style

            # Set styles for data rows
            alternate_fill_1 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            alternate_fill_2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column), 3):
                fill = alternate_fill_1 if row_idx % 2 == 0 else alternate_fill_2
                for cell in row:
                    cell.alignment = Alignment(horizontal="left")
                    cell.fill = fill
                    cell.border = border_style

            # Adjust column widths
            for col_idx, column_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, max_col=ws.max_column), 1):
                length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = length + 2

        # Save the Excel file
        wb.save(excel_path)

        print(f"Data successfully exported to '{excel_path}'.")

    except Error as e:
        print(f"Error connecting to or querying the database: {e}")

    except Exception as ex:
        print(f"An unexpected error occurred: {ex}")





#Informnation
mechanics_operation = """
The mechanic shop Kinich operates by receiving trucks that require various types of maintenance services. When a truck arrives at one of the shop's branches, the process begins with the manager creating an entrance order. This order is a critical document that captures all the essential information needed to service the truck effectively.
To facilitate this process, a suite of specialized tools has been incorporated into the shop's system. These tools streamline the creation and management of entrance orders, truck records, and product usage, ensuring that all data is accurately recorded and easily accessible.
Tools and Their Descriptions:
Create Camion Tool:
Requirements: VIN (Vehicle Identification Number), unit number, entry mileage, brand, and model.
Use Case: Used when a truck that is not yet in the system arrives at the shop. The manager inputs all necessary details about the truck to add it to the database, ensuring accurate maintenance records and future service planning.
Read Camion Tool:
Requirements: Truck's VIN.
Use Case: Allows staff to retrieve detailed information about a specific truck, ensuring the correct vehicle is associated with entrance orders and maintenance records.
Update Camion Tool:
Requirements: Truck's VIN and any details to be updated (unit number, mileage, brand, model).
Use Case: Used to update the truck's information in the database when changes occur, such as after servicing or correcting initial entry errors.
Delete Camion Tool:
Requirements: Truck's VIN.
Use Case: Removes a truck from the system, for example, if it is no longer in service or was entered incorrectly.
Create Order :
Requirements:  entry date, status, truck's VIN, maintenance reason,type of maintenance, branch location.
Use Case: Used by the manager to create an entrance order when a truck arrives for service. This tool captures all essential order details and interactively queries the user to associate products and consumables used during the service.
Read Orden Entrada Tool:
Requirements: Order ID.
Use Case: Allows staff to view and verify the details of entrance orders, including associated products and services.
Update Orden Entrada Tool:
Requirements: Order ID and any details to be updated (status, additional services, etc.).
Use Case: Used to update the entrance order with new information, such as status changes or additional required maintenance.
Delete Orden Entrada Tool:
Requirements: Order ID.
Use Case: Removes an entrance order from the system, for instance, if it was created in error.
Salida Orden Entrada Tool:
Requirements: Order ID, exit date,exit motive, status (should be updated to "liberada").
Use Case: Used by the manager to finalize an entrance order when service is completed. It records the exit date and changes the status to "liberada," indicating that the truck is ready for pickup.
Read Producto Tool:
Requirements: Product ID.
Use Case: Retrieves information about a specific product or consumable used in services.
Update Producto Tool:
Requirements: Product ID and any details to be updated (name, category).
Use Case: Used to update product information in the database, such as correcting details or updating categories.
Delete Producto Tool:
Requirements: Product ID.
Use Case: Removes a product from the system, for example, if it has been discontinued.
Read Productos Servicio Tool:
Requirements: Order ID.
Use Case: Provides a detailed list of all products and consumables associated with a specific entrance order, aiding in inventory tracking and billing.
Get Products Used in Month Tool:
Requirements: Year, month.
Use case: Retrieves all products used and their quantities in all orders during a specific month. This helps monitor product usage trends and plan inventory replenishment.
Get Product Usage for Truck Tool:
Requirements: Truck VIN, product ID, start date, end date.
Use case: Retrieves the total quantity of a specific product used for a particular truck within a time range. Useful for detecting anomalies in product usage, which may indicate misuse or theft.
Get Products Used for Truck Tool:
Requirements: Truck VIN.
Use case: Retrieves all products used and their quantities for a specific truck in all orders. Helps review the maintenance history and verify product usage for that truck.
Get Order Count for Branch Tool:
Requirements: Branch name or identifier.
Use case: Retrieves the number of orders processed by a specific branch. This helps assess branch performance and manage workload distribution.
Get Order Details for Truck Tool:
Requirements: Truck VIN.
Use case: Retrieves the number of orders, reasons, and dates for a specific truck. Provides a complete view of the truck's service history, assisting in maintenance planning and communication with the customer.
Operation Details:
Entrance Order Creation:
Establishes a relationship with the specific shop branch.
Includes detailed truck information: unique identification (VIN) and odometer reading at entry.
Documents the manager responsible for the order.
Specifies the maintenance type:
Corrective: Addressing existing issues or malfunctions.
Preventive: Performing routine checks and services to prevent future problems.
Consumable: Replacing regularly depleted items, such as oil or filters.
Product and Consumable Tracking:
The entrance order is linked to a table listing all products and consumables used.
Vital for inventory management, ensuring all parts and materials are accounted for.
Order Status Management:
Each order has a status reflecting its current stage:
Proceso: The truck is currently being serviced.
Inactiva: The service is temporarily halted.
Liberada: The service is completed, and the truck is ready for pickup.
Mechanics update the entrance order with any additional information or status changes during maintenance.
Service Completion:
Upon completion, the manager registers the exit date and changes the status to "liberada."
Indicates the truck has been fully serviced and is ready to be returned to the customer.
By meticulously recording all this information and utilizing the integrated tools, the mechanic shop ensures a seamless operation that tracks every aspect of the service process. This comprehensive system allows for:
Efficient Workflow Management: Tools like the Create Order with Products Tool and Update Orden Entrada Tool streamline the service workflow, reducing errors and saving time.
Accurate Billing: Linking products and services directly to entrance orders enables precise invoicing based on actual materials used and work performed.
Inventory Control: Associating products with orders helps monitor stock levels and plan reorders.
Data for Future Reference: Detailed service records support future diagnostics and service planning.
Enhanced Customer Satisfaction: Ensures transparency and accountability throughout the maintenance service.
Guidelines for Assistance:
Data Collection: Always gather all required information before proceeding with any operation.
Data Validation: Ensure the accuracy of the provided data.
Maintain Data Integrity: Follow the mechanic shop's procedures and ensure data consistency.
Communication: Provide clear and precise responses to support efficient workshop management.
Language Preference: Always respond in Spanish.
"""

categories_and_products = [
    {
        "category": "Electrico",
        "products": [
            "Plafon",
            "Encendedor",
            "Luces",
            "Fallo electrico",
            "Corto",
            "Modulos",
            "Control de arranque",
            "Modulo de bomba",
            "Sensores",
            "Problemas en transmisión",
            "Adaptador de poste",
            "Modulo PSM",
            "Acumuladores tornillo",
            "Líneas traseras",
            "Calavera",
            "Sincronizador y alternador",
            "Juego de cornetas",
            "Pines",
            "Botonera",
            "Marcha",
            "Convertidor",
            "Arnes",
            "Cilindro e interruptor",
            "Bulbo",
            "Selenoide",
            "Sistema ABS",
            "Actualizacion de calibracion de modulo ECM"
        ]
    },
    {
        "category": "Motor",
        "products": [
            "MP",
            "Refrigerador de gases",
            "Códigos",
            "Cambio de catalizador",
            "Cambio de polea",
            "Filtro",
            "Batería",
            "Deja de acelerar",
            "Falla al encender",
            "Pérdidas de potencia",
            "Cambio de clutch",
            "Embrague",
            "Daño en Compresor",
            "Abrazaderas",
            "Radiador",
            "Bomba de agua",
            "Secador de aire",
            "Válvula e inyector",
            "Regeneración de postramiento",
            "Dosificador de urea",
            "Escape",
            "Banda",
            "Ajuste de velocidad",
            "Turbo",
            "Cambio de parámetro",
            "Bayoneta"
        ]
    },
    {
        "category": "Carrocería",
        "products": [
            "Patines",
            "Chasis",
            "Camas de madera",
            "Faldón",
            "Caja seca",
            "Puertas y herrajes",
            "Concha",
            "Abrazaderas",
            "Esquinero",
            "Cóncavo",
            "Escalón",
            "Fascia",
            "Faro",
            "Cofre",
            "Láminas",
            "Filtración",
            "Extensión de defensa",
            "Gasket",
            "Zócalo",
            "Defensa",
            "Lodera",
            "Estribo",
            "Parrilla",
            "Tablero",
            "Rampas",
            "Alerón"
        ]
    },
    {
        "category": "Suspensión",
        "products": [
            "Balero",
            "Amortiguadores",
            "Masa delantera",
            "Muelle rota/dañada",
            "Percha",
            "Pernos",
            "Horquilla",
            "Crucetas y baleros",
            "Soporte aleta de tiburón",
            "Terminal y brazo viajero",
            "Buje",
            "Albardones",
            "Se engrasa la quinta",
            "Barra estabilizadora"
        ]
    },
    {
        "category": "Frenos",
        "products": [
            "Reemplazo en cilindro esclavo",
            "Frenos",
            "Rotochamber",
            "Balatas",
            "Mangueras de aire",
            "Bolsa de aire",
            "Discos traseros",
            "Matraca",
            "Mangueras hidráulicas",
            "Fuga de aire",
            "Mangueras servicio"
        ]
    },
    {
        "category": "Tren motriz",
        "products": [
            "Chicote",
            "Sincronizador doble de 3A",
            "Palanca de velocidades",
            "Resorte de presión corona",
            "Gases",
            "Flecha",
            "Diferencial",
            "Cardán",
            "Transmisión Clutch"
        ]
    },
    {
        "category": "Dirección",
        "products": [
            "Llantas nuevas",
            "Birlos",
            "RIN",
            "Colocación de llantas",
            "Colocación de rin",
            "Talachas",
            "Alineación",
            "Balanceo"
        ]
    },
    {
        "category": "Fluidos",
        "products": [
            "Relleno de niveles",
            "Aceite",
            "Urea",
            "Anticongelante"
        ]
    }
]


import sqlite3
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

def eliminar_duplicados(connection):
    cursor = connection.cursor()
    try:
        delete_query = """
        DELETE p1 FROM Productos p1
        INNER JOIN Productos p2 
        WHERE 
            p1.ID > p2.ID AND 
            p1.Nombre = p2.Nombre AND 
            p1.Categoria = p2.Categoria
        """
        cursor.execute(delete_query)
        connection.commit()
        print("Duplicated records deleted from Productos table.")
    except Error as e:
        print(f"An error occurred while deleting duplicates: {e}")
    finally:
        cursor.close()

def listar_productos_pdf(connection, output_pdf):
    try:
        cursor = connection.cursor()
        # Modificamos la consulta para no seleccionar el ID de la base de datos
        query = "SELECT Nombre, Categoria FROM Productos"
        cursor.execute(query)
        productos = cursor.fetchall()

        # Configuración del PDF con márgenes de 1 cm
        margin_size = 28.35  # 1 cm en puntos
        doc = SimpleDocTemplate(output_pdf, pagesize=letter,
                                rightMargin=margin_size, leftMargin=margin_size,
                                topMargin=margin_size, bottomMargin=margin_size)
        elements = []

        # Estilo de título
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.fontSize = 24
        title_style.leading = 28
        title_style.textColor = colors.black

        # Añadir el título en la primera página
        title = Paragraph("Lista de Productos y Servicios Kinich Multiservicios", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))  # Espacio entre título y tabla

        # Encabezado de la tabla
        data = [["#", "Nombre", "Categoría"]]  # Cambiamos 'ID' por '#' para numeración secuencial

        # Añadir productos a la data de la tabla con numeración secuencial
        for idx, producto in enumerate(productos, start=1):
            data.append([idx, producto[0], producto[1]])

        # Crear la tabla
        table = Table(data, colWidths=[1 * inch, 2.5 * inch, 2.5 * inch])

        # Aplicar estilos a la tabla
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Fondo gris para cabecera
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto blanco para cabecera
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Fondo beige para filas
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Borde de celda
        ])
        table.setStyle(style)

        # Añadir la tabla al documento
        elements.append(table)

        # Generar el PDF
        doc.build(elements)
        print(f"PDF generado con éxito en '{output_pdf}'.")

    except Error as e:
        print("Error al listar productos:", e)
    finally:
        cursor.close()




# Crear la base de datos y las tablas
initialize_database_mysql()

# Conectar a la base de datos MySQL
connection = mysql.connector.connect(
    host='db',  # Cambia si es necesario
    user='mecanicos_user',  # Reemplaza con tu usuario de MySQL
    password='mecanicos_password',  # Reemplaza con tu contraseña de MySQL
    database='mecanicos'
)

# Limpia la tabla antes de insertar datos
clear_productos_table(connection)

# Añadir categorías y productos
add_categories_and_products(connection, categories_and_products)

# Eliminar duplicados antes de generar el PDF
eliminar_duplicados(connection)

# Opcional: listar productos en PDF
#listar_productos_pdf(connection, "productos_lista.pdf")

# Cerrar la conexión
connection.close()